import openpyxl
import Aisuluu_bank
import pandas as pd

xl=openpyxl.load_workbook('bank.xlsx')



sheet=xl['Credit']
sheet_bank_account=xl['Bank account']
sheet_transfer=xl['Transfer history']

def client_menu(account_name):
    print('Hello client. \nWrite the number to choose.')
    while True:
        xl=openpyxl.load_workbook('bank.xlsx')
        sheet_bank_account=xl['Bank account']
        
        x=int(input('1-MyCredit   2-show Balance  3-to Convert  \n4-to Transfer  5-show Transfers History  6-Get Credit  7-Deposit  8-Exit  :'))
        
        if x==1:
            sheet=xl['Credit']
            for i in sheet['A']:
                if account_name in str(i.value):
                    account_num=i.row
            print('Credit sum = {0}som     Date = {1}      Percent = {2}%\n'.format(sheet['B'+str(account_num)].value,sheet['C'+str(account_num)].value,sheet['D'+str(account_num)].value))
        
        elif x==2:
            
            for i in sheet_bank_account['A']:
                if account_name in str(i.value):
                    account_num=i.row
            print('\nDollars = {}     Soms = {}       Rubs = {}       Euros = {}\n'.format(sheet_bank_account['B'+str(account_num)].value,sheet_bank_account['C'+str(account_num)].value,sheet_bank_account['D'+str(account_num)].value,sheet_bank_account['E'+str(account_num)].value))
        
        elif x==3:

            Aisuluu_bank.convert(account_name)
        
        elif x==4:

            Aisuluu_bank.transfer(account_name)

        elif x==5:
            xl=openpyxl.load_workbook('bank.xlsx')
            sheet_bank_account=xl['Bank account']
            sheet_transfer=xl['Transfer history']
            for i in sheet_transfer['A']:
                if account_name in str(i.value):
                    account_num=i.row
            print('From me')
            print(sheet_transfer['B'+str(account_num)].value)
            print('To me')
            print(sheet_transfer['C'+str(account_num)].value)
        elif x==6: 
            Aisuluu_bank.get_credit(account_name)

        elif x==8:
            print('Exit...\nThank you for using our services.')
            return 6
        elif x==7: 
            sheet=xl['Credit']
            sheet_bank_account=xl['Bank account']
            money=int(input('how much: ')) 
            for i in sheet['A']: 
                if account_name in str(i.value): 
                    account_num=i.row 
            a=int(input('what currency?\n1-doll     2-som       3-rub       4-euro: '))   
            if a==1: 
                sheet_bank_account['B'+str(account_num)].value+=money 
            elif a==2: 
                sheet_bank_account['C'+str(account_num)].value+=money 
            elif a==3: 
                sheet_bank_account['D'+str(account_num)].value+=money 
            elif a==4: 
                sheet_bank_account['E'+str(account_num)].value+=money
            xl.save('bank.xlsx')
        else:
            break

# client_menu('jojoman@gmail.com')