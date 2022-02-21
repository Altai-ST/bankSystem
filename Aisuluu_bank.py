import openpyxl
from datetime import datetime, timedelta, date
import math
import pandas as pd
from dateutil.relativedelta import relativedelta
x2=openpyxl.load_workbook('valet.xlsx')
xl=openpyxl.load_workbook('bank.xlsx')
sheet_bank_account=xl['Bank account']
sheet_currency=x2['Kurs']
sheet_prices=xl['Prices']
sheet=xl['Credit']
sheet_transfer=xl['Transfer history']
sheet_credit=xl['Credit Setting']
# 1
book = openpyxl.load_workbook('./bank.xlsx')

def add_user_credit(url, user_data):
    with pd.ExcelWriter(url, engine = 'openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        user_data.to_excel(writer, 'Credit', index=False)
        writer.save()

def read_user(url, sheet):
    read = pd.read_excel(url, sheet_name=sheet, index_col=False)
    return read

def convert(account_name):

    for i in sheet_bank_account['A']:
            if account_name in str(i.value):
                account_num=i.row

    dollar=float(sheet_currency['A2'].value)
    som=float(sheet_currency['B2'].value)
    rub=float(sheet_currency['D2'].value)
    euro=float(sheet_currency['C2'].value)

    dollmoney=float(sheet_bank_account['B'+str(account_num)].value)
    sommoney=float(sheet_bank_account['C'+str(account_num)].value)
    rubmoney=float(sheet_bank_account['D'+str(account_num)].value)
    euromoney=float(sheet_bank_account['E'+str(account_num)].value)


    a=int(input('From: 1-dollas, 2-som, 3-rub, 4-euro =='))
    if a==1:
        a=dollar;money=dollmoney;name='doll'
    elif a==2:
        a=som;money=sommoney;name='som'
    elif a==3:
        a=rub;money=rubmoney;name='rub'
    elif a==4:
        a=euro;money=euromoney;name='euro'

    b=int(input('To: 1-dollas, 2-som, 3-rub, 4-euro =='))
    if b==1:
        b=dollar;name2='doll'
    elif b==2:
        b=som;name2='som'
    elif b==3:
        b=rub;name2='rub'
    elif b==4:
        b=euro;name2='euro'

    charge=float(sheet_prices['B1'].value)
    
    user_credit_sum=int(input('How much?  =='))
    while user_credit_sum > money and user_credit_sum <0:
        print("You don't have enough.")
        user_credit_sum=float(input('How much?  =='))



    # convert result
    result=round(user_credit_sum/(b*(1/a)),2)
    
    # percents
    service=round(result*charge,2)

    # showing results
    if b==dollar and name2=='doll':
        print('result is: {} doll'.format(result))
    elif b==som and name2=='som':
        print('result is: {} som'.format(result))
    elif b==rub and name2=='rub':
        print('result is:  {} rub'.format(result))
    elif b==euro and name2=='euro':
        print('result is: {} euro'.format(result))

    # showing percents
    if a==dollar and name=='doll':
        print('service -%: {} doll'.format(service))
    elif a==som and name=='doll':
        print('service -%: {} som'.format(service))
    elif a==rub and name=='doll':
        print('service -%: {} rub'.format(service))
    elif a==euro and name=='doll':
        print('service -%: {} euro'.format(service))


    # + in balance
    if b==dollar:
        dollmoney+=result-service
    elif b==som:
        sommoney+=result-service
    elif b==rub:
        rubmoney+=result-service
    elif b==euro:
        euromoney+=result-service

    sheet_bank_account['B'+str(account_num)].value=round(dollmoney,2)
    sheet_bank_account['C'+str(account_num)].value=round(sommoney,2)
    sheet_bank_account['D'+str(account_num)].value=round(rubmoney,2)
    sheet_bank_account['E'+str(account_num)].value=round(euromoney,2)
    xl.save('bank.xlsx')

    # - from balance
    if a==dollar:
        dollmoney-=user_credit_sum
    elif a==som:
        sommoney-=user_credit_sum
    elif a==rub:
        rubmoney-=user_credit_sum
    elif a==euro:
        euromoney-=user_credit_sum
    
    sheet_bank_account['B'+str(account_num)].value=round(dollmoney,2)
    sheet_bank_account['C'+str(account_num)].value=round(sommoney,2)
    sheet_bank_account['D'+str(account_num)].value=round(rubmoney,2)
    sheet_bank_account['E'+str(account_num)].value=round(euromoney,2)
    xl.save('bank.xlsx')

# 2
def transfer(account_name):
    charge=float(sheet_prices['B2'].value)

    for i in sheet_bank_account['A']:
            if account_name in str(i.value):
                account_num=i.row
    
    def read_user(url, sheet):
        read = pd.read_excel(url, sheet_name=sheet, index_col=False)
        return read

    # show accounts
    for i in sheet_bank_account['A']:              
            if str(i.value) == account_name:
                print(str(i.value),'(you)')

    # to whom
    users = read_user('./account.xlsx', 'User')['Email']
    print(users)
    to=input('Enter whom you want to transfer    ==')
    
    # searches account
    for i in sheet_bank_account['A']:
            if to in str(i.value):
                second_account_num=i.row


    # currency
    a=int(input('From: 1-dollar, 2-som, 3-rub, 4-euro   =='))
    if a==1:
        money=float(sheet_bank_account['B'+str(account_num)].value)
    elif a==2:
        money=float(sheet_bank_account['C'+str(account_num)].value)
    elif a==3:
        money=float(sheet_bank_account['D'+str(account_num)].value)
    elif a==4:
        money=float(sheet_bank_account['E'+str(account_num)].value)

    # how much
    user_credit_sum=float(input('How much?  =='))
    while user_credit_sum>money:
        print("You don't have enough.")
        user_credit_sum=float(input('How much?  =='))
        continue

    service=user_credit_sum*float(sheet_prices['B1'].value)

    # + in balance of receiver
    if a==1:
        sheet_bank_account['B'+str(second_account_num)].value=round(sheet_bank_account['B'+str(second_account_num)].value+user_credit_sum-service,2)
    elif a==2:
        sheet_bank_account['C'+str(second_account_num)].value=round(sheet_bank_account['C'+str(second_account_num)].value+user_credit_sum-service,2)
    elif a==3:
        sheet_bank_account['D'+str(second_account_num)].value=round(sheet_bank_account['D'+str(second_account_num)].value+user_credit_sum-service,2)
    elif a==4:
        sheet_bank_account['E'+str(second_account_num)].value=round(sheet_bank_account['E'+str(second_account_num)].value+user_credit_sum-service,2)
    # - from own balance
    if a==1:
        sheet_bank_account['B'+str(account_num)].value=round(sheet_bank_account['B'+str(account_num)].value-user_credit_sum,2)
    elif a==2:
        sheet_bank_account['C'+str(account_num)].value=round(sheet_bank_account['C'+str(account_num)].value-user_credit_sum,2)
    elif a==3:
        sheet_bank_account['D'+str(account_num)].value=round(sheet_bank_account['D'+str(account_num)].value-user_credit_sum,2)
    elif a==4:
        sheet_bank_account['E'+str(account_num)].value=round(sheet_bank_account['E'+str(account_num)].value-user_credit_sum,2)
    
    # saving history note
    # history
    if a==1:
        history='{}\n-{} sends {}doll to {}'.format(sheet_transfer['B'+str(account_num)].value,account_name,user_credit_sum,to)
        history2='{}\n-{} gets {}doll from {}'.format(sheet_transfer['B'+str(account_num)].value,to,user_credit_sum,account_name)
    elif a==2:
        history='{}\n-{} sends {}som to {}'.format(sheet_transfer['B'+str(account_num)].value,account_name,user_credit_sum,to)
        history2='{}\n-{} gets {}som from {}'.format(sheet_transfer['B'+str(account_num)].value,to,user_credit_sum,account_name)
    elif a==3:
        history='{}\n-{} sends {}rub to {}'.format(sheet_transfer['B'+str(account_num)].value,account_name,user_credit_sum,to)
        history2='{}\n-{} gets {}rub from {}'.format(sheet_transfer['B'+str(account_num)].value,to,user_credit_sum,account_name)
    elif a==4:
        history='{}\n-{} sends {}euro to {}'.format(sheet_transfer['B'+str(account_num)].value,account_name,user_credit_sum,to)
        history2='{}\n-{} gets {}euro from {}'.format(sheet_transfer['B'+str(account_num)].value,to,user_credit_sum,account_name)
    print('-{} sends {}doll to {}'.format(account_name,user_credit_sum,to))
    sheet_transfer['B'+str(account_num)].value=history
    sheet_transfer['C'+str(second_account_num)].value=history2
    xl.save('bank.xlsx')


def credit_setting():
    import openpyxl
    xl=openpyxl.load_workbook('bank.xlsx')
    sheet_credit=xl['Credit Setting']

    def set_credit_setting(text, position):
        set_credit=int(input(text))
        sheet_credit[position].value=set_credit
        xl.save('bank.xlsx')

    credit_setting_menu=int(input('Do you want to change:\n   1 - Maximal credit ({})\n   2 - Minimal credit ({})\n   3 - Percent rate ({})\n   4 - Credit date ({})\n:'.format(sheet_credit['B2'].value,sheet_credit['B1'].value,sheet_credit['B3'].value,sheet_credit['B4'].value)))
    if credit_setting_menu==1:
        set_credit_setting('New maximal credit:','B2')
    elif credit_setting_menu==2:
        set_credit_setting('New minimal credit:','B1')
    elif credit_setting_menu==3:
        set_credit_setting('New percent rate:','B3')
    elif credit_setting_menu==4:
        set_credit_setting('Input month:','B4')

# credit
def get_credit(account_name):
    
    xl=openpyxl.load_workbook('bank.xlsx')
    sheet_credit=xl['Credit Setting']
    sheet=xl['Credit']
    credit_menu=int(input('Do you want a Credit in the national currency - som\n1-yes 0-no\n: '))

    # if sheet['A2'] == None:
    #     print('hy')
    #     all_credits = read_user('./bank.xlsx', 'Credit') 
    #     local_credits = pd.DataFrame(all_credits)
    #     new_user_credit = local_credits.append({'email': account_name, 'Credit': None, 'Date': None, 'Percent': None}, ignore_index=True)
    #     add_user_credit('./bank.xlsx', new_user_credit)

    credit_info = read_user('./bank.xlsx', 'Credit')
    if account_name not in credit_info['email'].values:
        all_credits = read_user('./bank.xlsx', 'Credit') 
        local_credits = pd.DataFrame(all_credits)
        new_user_credit = local_credits.append({'email': account_name, 'Credit': None, 'Date': None, 'Percent': None}, ignore_index=True)
        add_user_credit('./bank.xlsx', new_user_credit)


    if credit_menu == 1:
        max_credit = sheet_credit['B2'].value
        min_credit = sheet_credit['B1'].value
        rate = sheet_credit['B3'].value

        for i in sheet['A']:
                if account_name in str(i.value):
                    account_num=i.row
        user_credit_sum=int(input('Maximal credit ({})  Minimal credit ({})\nHow much?  =='.format(max_credit,sheet_credit['B1'].value)))
        while user_credit_sum>max_credit or user_credit_sum<min_credit:
            if user_credit_sum>max_credit:
                print('You cant recive so much')
            elif user_credit_sum<min_credit:
                print("That's not enough")
            user_credit_sum=int(input('Maximal credit ({})  Minimal credit ({})\nHow much?  =='.format(max_credit,sheet_credit['B1'].value)))
            

        sheet_bank_account['C'+str(account_num)].value+=user_credit_sum 
        print(sheet_bank_account['C'+str(account_num)].value)

        xl.save('bank.xlsx')


        if sheet['B'+str(account_num)].value == None: 
            sheet['B'+str(account_num)].value=user_credit_sum 
        else:  
            sheet['B'+str(account_num)].value+=user_credit_sum
        sheet['D'+str(account_num)].value=rate
        sheet['C'+str(account_num)].value=datetime.now().date()+relativedelta(months=sheet_credit['B4'].value)
        
        xl.save('bank.xlsx')

        print('You succesfully got {}som credit'.format(user_credit_sum))
    else:
        return