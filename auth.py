import BankMenu as bm
from openpyxl import load_workbook
import pandas as pd
import base

book = load_workbook('./account.xlsx')
def add_user(url, user_data):
    with pd.ExcelWriter(url, engine = 'openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        user_data.to_excel(writer, 'User', index=False)
        writer.save()

book_cost = load_workbook('./bank.xlsx')
def add_user_cost(url, user_data, second_data, history_data):
    with pd.ExcelWriter(url, engine = 'openpyxl') as writer:
        writer.book = book_cost
        writer.sheets = dict((ws.title, ws) for ws in book_cost.worksheets)
        user_data.to_excel(writer, 'Bank account', index=False)
        second_data.to_excel(writer, 'Credit', index=False)
        history_data.to_excel(writer, 'Transfer history', index=False)
        writer.save()

def read_user(url, sheet):
    read = pd.read_excel(url, sheet_name=sheet, index_col=False)
    return read


def sign_in():
    while True:
        print('\t\t\t\tSign in')
        sign_in_email = input('Enter an email: ')
        sign_in_password = input('Enter a password: ')
        print('1. Sign in')
        print('2. Cancel')
        choose_sign_in = int(input(':'))
        if choose_sign_in == 1:
            exit = 0
            isWorker = read_user('./account.xlsx', 'Worker')
            for i in range(len(isWorker['Email'])):
                if (sign_in_email == isWorker['Email'][i]) and (sign_in_password == isWorker['Password'][i]):
                    print('\t\t\t\t\tWelcome ',isWorker['Name'][i],' !')
                    exit = bm.worker_menu()
            isUser = read_user('./account.xlsx', 'User')
            for i in range(len(isUser['Email'])):
                if (sign_in_email == isUser['Email'][i]) and (sign_in_password == isUser['Password'][i]):
                    print('\t\t\t\t\tWelcome ',isUser['Name'][i],' !')
                    exit = base.client_menu(isUser['Email'][i])
            if exit == 6:
                break
        elif choose_sign_in == 2:
            registration()
        print("User not found")
        print('Do it again')


def registration():
    while True:
        print('\t\t\t\tWelcome to the Aisuluu bank')
        print('Select login type:\n\t 1.Sign in \n\t 2.Registration')
        auth_choose = int(input(':'))
        if auth_choose == 2:
            print('\t\t\t\tRegistration:')
            sign_up_name = input('Enter a name: ')
            sign_up_email = input('Enter an email: ')
            sign_up_password = input('Enter a password: ')

            if len(sign_up_password) >= 8 and ('@gmail.com' in sign_up_email or '@mail.ru' in sign_up_email) and sign_up_name != '':
                saved_user = read_user('./account.xlsx', 'User')
                local_user = pd.DataFrame(saved_user)
                update_user = local_user.append({'Name': sign_up_name, 'Email': sign_up_email, 'Password': sign_up_password}, ignore_index=True)
                add_user('./account.xlsx', update_user)

                saved_user_cost = read_user('./bank.xlsx', 'Bank account')
                local_user_cost = pd.DataFrame(saved_user_cost)
                update_user_cost = local_user_cost.append({'email': sign_up_email, 'Doll': 0, 'Som': 0, 'Rub':0, 'Eur': 0}, ignore_index=True)

                saved_user_credit = read_user('./bank.xlsx', 'Credit')
                local_user_credit = pd.DataFrame(saved_user_credit)
                update_user_credit = local_user_credit.append({'email': sign_up_email, 'Credit': None, 'Date': None, 'Percent': None }, ignore_index=True)
                 
                saved_user_history = read_user('./bank.xlsx', 'Transfer history')
                local_user_history = pd.DataFrame(saved_user_history)
                update_user_history = local_user_history.append({'email': sign_up_email, 'give': '', 'get': ''}, ignore_index=True)

                add_user_cost('./bank.xlsx', update_user_cost, update_user_credit, update_user_history)

                print('Congratulations on successfully registering!!!')
                sign_in()

            elif len(sign_up_password) < 8:
                print('The password must be greater than 8')
            elif not(('@' in sign_up_email) and ('.com' in sign_up_email)):
                print('Wrong email address')
        elif auth_choose == 1:
            sign_in()
        
registration()