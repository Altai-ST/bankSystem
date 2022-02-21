import pandas as pd
from openpyxl import load_workbook

# kurs_valut = pd.DataFrame({'Dollar':[1], 'Som':[1], 'Euro':[1], 'Rub':[1]})

# kurs_valut.to_excel('./valet.xlsx', sheet_name='Kurs', index=False, engine='openpyxl')


def read_user(url, sheet):
    read = pd.read_excel(url, sheet_name=sheet, index_col=False)
    return read

book = load_workbook('./valet.xlsx')
def add_valet(url, valet_data):
    with pd.ExcelWriter(url, engine = 'openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        valet_data.to_excel(writer, 'Kurs', index=False)
        writer.save()

def update_valet(new_kurs, valet):
    valet_data = read_user('./valet.xlsx', 'Kurs')
    valet_data[valet][0] = new_kurs
    print('val: ',valet_data[valet])
    local_valet = pd.DataFrame(valet_data)
    add_valet('./valet.xlsx', local_valet)

def setting_convert():
    while True:
        print('\t\t\t\t\t\tSet a new currency rate')
        print('1.Set a new currency exchange rate for the dollar\n2.Set a new currency exchange rate for the som\n3.Set a new currency exchange rate for the euro')
        print('4.Set a new currency exchange rate for the ruble\n5.Back to the main menu')
        
        choose_set_valet = int(input(':'))
        if choose_set_valet == 1:
            print('\t\t\t\t\tChange the currency rate')
            new_kurs_currency = input('Dollar:')
            print('sum:', new_kurs_currency)
            update_valet(new_kurs_currency, 'Dollar')
        elif choose_set_valet == 2:
            print('\t\t\t\t\tChange the currency rate')
            new_kurs_currency = input('Som:')
            update_valet(new_kurs_currency, 'Som')
        elif choose_set_valet == 3:
            print('\t\t\t\t\tChange the currency rate')
            new_kurs_currency = input('Euro:')
            update_valet(new_kurs_currency, 'Euro')
        elif choose_set_valet == 4:
            print('\t\t\t\t\tChange the currency rate')
            new_kurs_currency = input('Rub:')
            update_valet(new_kurs_currency, 'Rub')
        elif choose_set_valet == 5:
            break

def credit_setting():
    import openpyxl
    xl=openpyxl.load_workbook('bank.xlsx')
    sheet_credit=xl['Credit Setting']

    def set_credit_setting(text, position):
        set_credit=int(input(text))
        sheet_credit[position].value=set_credit
        xl.save('bank.xlsx')

    credit_setting_menu=int(input('Do you want to change:\n   1 - Maximal credit ({})\n   2 - Minimal credit ({})\n   3 - Percent rate ({})\n   4 - Credit date ({})\n:'.format(sheet_credit['B2'].value,sheet_credit['B1'].value,sheet_credit['B3'].value,sheet_credit['B4'].value)))
    if credit_setting_menu==1:
        set_credit_setting('New maximal credit:','B2')
    elif credit_setting_menu==2:
        set_credit_setting('New minimal credit:','B1')
    elif credit_setting_menu==3:
        set_credit_setting('New percent rate:','B3')
    elif credit_setting_menu==4:
        set_credit_setting('Input month:','B4')
